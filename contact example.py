import openpyxl
import sys
"""
If you want to use the "sys.exit()" to terminate the process 
automatically, you should import sys
"""

import subprocess
#A convinient lib to execute windows applications with python

book = openpyxl.load_workbook('./*.xlsx') 
#change the ./* to your excel file address
sheet = book["Sheet1"]

ans = ["name", "company", "title", "tel", "mail"]
#You can add more content if you want

p = " plz:"
celnum = 4

while True:
	l = sheet.cell(row = celnum, column = 3).value
	#In most cases, peolple save some space to start forming

	if l == None:
		#Find the first cel that has no value written
		print(ans[0] + p)
		name = input()
		#Write in your information
		
		if name == 'end':
		#When you type "end" you can kill the process	
			
			book.save('./*.xlsx')
			subprocess.Popen([r'./office/excel.exe', './*.xlsx'])
			"""
			As same as the previous comment, put your file address
			int the blank
			"""

			sys.exit()
		
		else:
			sheet.cell(row = celnum, column = 3).value = name
			print(ans[1] + p)
			cpn = input()
			sheet.cell(row = celnum, column = 4).value = cpn
			print(ans[2] + p)
			ttl = input()
			sheet.cell(row = celnum, column = 5).value = ttl
			print(ans[3] + p)
			tel = input()
			sheet.cell(row = celnum, column = 6).value = tel
			print(ans[4] + p)
			ml = input()
			sheet.cell(row = celnum, column = 7).value = ml
			#Writing your information

	elif l != None:
		print("passed")
		celnum += 1
		#If the cel had been written, jump to the next line